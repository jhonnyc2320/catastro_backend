from rest_framework import serializers
import zipfile
import csv
from collections import defaultdict
import openpyxl
import xlrd
from rest_framework import generics
from apps.catastro.models import Asignacion, LcDcPredio as Predio

class CargueDatosSerializer(serializers.Serializer):
    files = serializers.ListField()

    def create(self, validated_data):
        return validated_data
    


class ReadExcelAsignacion(serializers.Serializer):
    file = serializers.FileField(write_only=True)

    def validate(self, attrs):
        file = attrs.get('file')
        if not self.is_excel_file(file):
            raise serializers.ValidationError("El archivo proporcionado no es un archivo Excel válido.")
        return attrs
    
    def is_excel_file(self, file):
        try:
            # Intenta cargar con openpyxl (para archivos .xlsx)
            openpyxl.load_workbook(file)
            return True
        except openpyxl.utils.exceptions.InvalidFileException:
            pass

        try:
            # Intenta cargar con xlrd (para archivos .xls)
            xlrd.open_workbook(file_contents=file.read())
            return True
        except xlrd.biffh.XLRDError:
            pass

        return False
    
    def to_representation(self, instance):
        file = instance.get('file')
        data = self.read_excel_file(file)
        return {"data": data}
    
    def read_excel_file(self, file):
        data = []
        try:
            # Intenta cargar con openpyxl (para archivos .xlsx)
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Asumiendo que la primera fila es el encabezado
                data.append({
                    'npn': row[0],
                    'coordinador': row[1]
                })
        except openpyxl.utils.exceptions.InvalidFileException:
            file.seek(0)  # Reiniciar el puntero del archivo
            # Intenta cargar con xlrd (para archivos .xls)
            wb = xlrd.open_workbook(file_contents=file.read())
            sheet = wb.sheet_by_index(0)
            for row_idx in range(1, sheet.nrows):  # Asumiendo que la primera fila es el encabezado
                row = sheet.row_values(row_idx)
                data.append({
                    'npn': row[0],
                    'coordinador': row[1]
                })
        return data


class ZipFileSerializer(serializers.Serializer):
    file = serializers.FileField(write_only=True)

    def validate(self, attrs):
        file = attrs.get('file')
        if not zipfile.is_zipfile(file):
            raise serializers.ValidationError("El archivo proporcionado no es un archivo .zip válido.")

        try:
            with zipfile.ZipFile(file, 'r') as zip_ref:
                zip_contents = zip_ref.namelist()

                if len(zip_contents) != 11 or not all(name.endswith('.csv') for name in zip_contents):
                    raise serializers.ValidationError("El archivo .zip debe contener exactamente 11 archivos .csv.")

                if 'idpredio.csv' not in zip_contents or zip_ref.getinfo('idpredio.csv').file_size == 0:
                    raise serializers.ValidationError("El archivo idpredio.csv está vacío o no está presente.")

                idpredio_data = self.read_csv(zip_ref, 'idpredio.csv')
                terreno_data = self.read_csv(zip_ref, 'terreno.csv')
                construccion_data = self.read_csv(zip_ref, 'construccion.csv')  # Leer datos de construcción
                fuente_admin_data = self.read_csv(zip_ref, 'fuenteadministrativa.csv')
                coordenadas_data = self.read_csv(zip_ref, 'coordenadas.csv')
                contacto_visita_data = self.read_csv(zip_ref, 'contactovisita.csv')
                interesado_predio_data = self.read_csv(zip_ref, 'interesadopredio.csv')
                interesado_data = self.read_csv(zip_ref, 'interesado.csv')
                
                return self.process_data(idpredio_data, terreno_data, construccion_data, fuente_admin_data, coordenadas_data, contacto_visita_data, interesado_predio_data, interesado_data)
        except Exception as e:
            raise serializers.ValidationError("Error procesando el archivo .zip: " + str(e))

    def create(self, validated_data):
        # Aquí se podría crear una instancia del modelo si es necesario
        return validated_data

    def process_data(self, idpredio_data, terreno_data, construccion_data, fuente_admin_data, coordenadas_data, contacto_visita_data, interesado_predio_data, interesado_data):
        data_by_id = defaultdict(lambda: {
            'terrenos': [],
            'construcciones': [],
            'fuentes_administrativas': [],
            'coordenadas': [],
            'contactos_visita': [],
            'interesados': []
        })

        # Organizar los datos de interesados e interesadopredio
        interesados_by_id = {interesado['ID_INTERESADO']: interesado for interesado in interesado_data}
        interesado_predio_by_predio = defaultdict(list)
        for ip in interesado_predio_data:
            if ip['ID_INTERESADO'] in interesados_by_id:
                interesado_info = interesados_by_id[ip['ID_INTERESADO']].copy()
                interesado_info['fuentes_administrativas'] = [fa for fa in fuente_admin_data if fa['ID_REGISTRO'] == ip['ID_REGISTRO']]
                interesado_predio_by_predio[ip['ID_PREDIO']].append(interesado_info)

        # Agrupar los datos de cada archivo por ID_PREDIO
        for dataset, key in [
            (terreno_data, 'terrenos'),
            (construccion_data, 'construcciones'),
            (fuente_admin_data, 'fuentes_administrativas'),
            (coordenadas_data, 'coordenadas'),
            (contacto_visita_data, 'contactos_visita')
        ]:
            for item in dataset:
                id_predio = item.get('ID_PREDIO')
                if id_predio:
                    data_by_id[id_predio][key].append({k.lower(): v for k, v in item.items()})

        # Combinar todos los datos con los de idpredio
        combined_data = []
        for predio in idpredio_data:
            predio_info = {k.lower(): v for k, v in predio.items()}
            predio_info.update(data_by_id.get(predio['ID_PREDIO'], {}))
            predio_info['interesados'] = interesado_predio_by_predio.get(predio['ID_PREDIO'], [])
            combined_data.append(predio_info)
        
        return combined_data

    def read_csv(self, zip_ref, filename):
        with zip_ref.open(filename) as file:
            reader = csv.DictReader((line.decode('utf-8') for line in file), delimiter='|')
            return list(reader)